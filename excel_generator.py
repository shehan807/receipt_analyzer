"""Excel report generation with formatting."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Category, ReceiptData


class ExcelGenerator:
    """Generate formatted Excel reports for receipt analysis."""

    # Color definitions
    COLORS = {
        "green": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "red": PatternFill(start_color="FFB6B6", end_color="FFB6B6", fill_type="solid"),
        "blue": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "orange": PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid"),
        "yellow": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
        "header": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "light_red": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    }

    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def generate(self, receipts: list[ReceiptData], output_path: Path, center_name: str = None) -> None:
        """Generate Excel report with summary and item details.

        Args:
            receipts: List of receipt data to include.
            output_path: Path for the output Excel file (directory or full path).
            center_name: Optional center name to include in filename.
        """
        # If center_name provided, update filename to summary_{center_name}.xlsx
        if center_name:
            # Sanitize center name for filename (remove special characters)
            safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in center_name)
            safe_name = safe_name.strip().replace(' ', '_')
            output_path = output_path.parent / f"summary_{safe_name}.xlsx"

        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, receipts)

        # Sheet 2: Item Details
        ws_items = wb.create_sheet("Item Details")
        self._create_items_sheet(ws_items, receipts)

        wb.save(output_path)

    def _create_summary_sheet(self, ws, receipts: list[ReceiptData]) -> None:
        """Create the summary sheet.

        Args:
            ws: Worksheet to populate.
            receipts: List of receipt data.
        """
        # Headers
        headers = [
            "Filename",
            "Store",
            "Date",
            "Subtotal",
            "Tax Paid?",
            "Tax Amt",
            "Delivery",
            "Total",
            "Payment",
            "Category",
            "Needs Review?",
            "Review Reasons",
        ]
        ws.append(headers)

        # Style headers
        for col, cell in enumerate(ws[1], 1):
            cell.fill = self.COLORS["header"]
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")

        # Sort: flagged items first, then by filename
        sorted_receipts = sorted(
            receipts, key=lambda r: (not r.needs_review, r.source_file)
        )

        for row_idx, receipt in enumerate(sorted_receipts, 2):
            # Format date
            date_str = ""
            if receipt.receipt_date:
                date_str = receipt.receipt_date.isoformat()
            elif receipt.date_raw:
                date_str = receipt.date_raw

            # Format payment info
            payment_str = ""
            if receipt.payment_method:
                payment_str = receipt.payment_method
                if receipt.card_last_four:
                    payment_str += f" ****{receipt.card_last_four}"

            ws.append(
                [
                    receipt.source_file,
                    receipt.store_name,
                    date_str,
                    receipt.subtotal if receipt.subtotal else "",
                    "Yes" if receipt.tax_paid else "No",
                    receipt.tax_amount if receipt.tax_amount else "",
                    receipt.delivery_cost if receipt.delivery_cost else "",
                    receipt.total,
                    payment_str,
                    receipt.overall_category.value.title(),
                    "Yes" if receipt.needs_review else "No",
                    "; ".join(receipt.review_reasons) if receipt.review_reasons else "",
                ]
            )

            # Apply formatting
            for col in range(1, 13):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.BORDER

            # Color Tax Paid column (column 5)
            tax_cell = ws.cell(row=row_idx, column=5)
            tax_cell.fill = self.COLORS["green"] if receipt.tax_paid else self.COLORS["red"]

            # Color Payment column (column 9) - red if no payment info
            payment_cell = ws.cell(row=row_idx, column=9)
            if not receipt.payment_method:
                payment_cell.fill = self.COLORS["red"]

            # Highlight entire row if needs review (except tax and payment columns)
            if receipt.needs_review:
                for col in range(1, 13):
                    if col not in (5, 9):  # Don't override tax/payment colors
                        ws.cell(row=row_idx, column=col).fill = self.COLORS["light_red"]

            # Color category column (column 10)
            cat_cell = ws.cell(row=row_idx, column=10)
            if receipt.overall_category == Category.SNACK:
                cat_cell.fill = self.COLORS["orange"]
            elif receipt.overall_category == Category.SUPPLY:
                cat_cell.fill = self.COLORS["blue"]
            elif receipt.overall_category in (Category.QUESTIONABLE, Category.UNKNOWN):
                cat_cell.fill = self.COLORS["yellow"]

        # Add totals row
        totals_row = len(sorted_receipts) + 2
        last_data_row = totals_row - 1
        ws.cell(row=totals_row, column=1, value="TOTALS")
        ws.cell(row=totals_row, column=1).font = Font(bold=True)

        # Use Excel SUM formulas for totals
        ws.cell(row=totals_row, column=4, value=f"=SUM(D2:D{last_data_row})")
        ws.cell(row=totals_row, column=6, value=f"=SUM(F2:F{last_data_row})")
        ws.cell(row=totals_row, column=7, value=f"=SUM(G2:G{last_data_row})")
        ws.cell(row=totals_row, column=8, value=f"=SUM(H2:H{last_data_row})")

        # Style totals row
        for col in range(1, 13):
            cell = ws.cell(row=totals_row, column=col)
            cell.font = Font(bold=True)
            cell.border = self.BORDER

        # Add auto-filter (excluding totals row)
        ws.auto_filter.ref = f"A1:L{len(sorted_receipts) + 1}"

        # Adjust column widths
        self._auto_adjust_columns(ws)

    def _create_items_sheet(self, ws, receipts: list[ReceiptData]) -> None:
        """Create the item details sheet.

        Args:
            ws: Worksheet to populate.
            receipts: List of receipt data.
        """
        # Headers
        headers = ["Filename", "Store", "Item", "Qty", "Price", "Category", "Confidence", "Verified"]
        ws.append(headers)

        # Style headers
        for col, cell in enumerate(ws[1], 1):
            cell.fill = self.COLORS["header"]
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")

        # Collect all items with their source
        all_items = []
        for receipt in receipts:
            for item in receipt.items:
                all_items.append((receipt.source_file, receipt.store_name, item))

        # Sort: questionable/unknown first, then by category, then by filename
        category_order = {
            Category.QUESTIONABLE: 0,
            Category.UNKNOWN: 0,
            Category.SNACK: 1,
            Category.SUPPLY: 2,
        }
        all_items.sort(
            key=lambda x: (category_order.get(x[2].category, 3), x[0], x[2].name)
        )

        for row_idx, (filename, store, item) in enumerate(all_items, 2):
            ws.append(
                [
                    filename,
                    store,
                    item.name,
                    item.quantity,
                    item.price,
                    item.category.value.title(),
                    item.confidence.value.title(),
                    "",  # Empty Verified column for human checkoff
                ]
            )

            # Apply formatting
            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.BORDER

            # Color by category (column 6)
            category_cell = ws.cell(row=row_idx, column=6)
            if item.category == Category.SUPPLY:
                category_cell.fill = self.COLORS["blue"]
            elif item.category == Category.SNACK:
                category_cell.fill = self.COLORS["orange"]
            elif item.category in (Category.QUESTIONABLE, Category.UNKNOWN):
                category_cell.fill = self.COLORS["yellow"]

        # Add auto-filter
        if all_items:
            ws.auto_filter.ref = f"A1:H{len(all_items) + 1}"

        # Adjust column widths
        self._auto_adjust_columns(ws)

    def _auto_adjust_columns(self, ws) -> None:
        """Auto-adjust column widths based on content.

        Args:
            ws: Worksheet to adjust.
        """
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except (TypeError, AttributeError):
                    pass

            # Set width with some padding, but cap at reasonable max
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
